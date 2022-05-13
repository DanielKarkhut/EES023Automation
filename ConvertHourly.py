# Built by Daniel Karkhut
# SPDX-License-Identifier: UNLICENSED

from openpyxl import load_workbook
from datetime import datetime
from datetime import date
from datetime import timedelta

# Day Data Structure
class DayData:
    def __init__(self, currDay):
        self.day = currDay
        self.first = -1
        self.second = -1
        self.third = -1
        self.fourth = -1

# Load Workbook
wb2 = load_workbook(filename = 'WeatherNotebook.xlsx')

# Sheet with all hourly recordings
dataSheet = wb2['Sheet1']

# Create sheet with 4 daily hourly recordings
cleanDataSheet = wb2.create_sheet('CleanData')

# Store Max Row and Max Column Sizes
maxr = dataSheet.max_row
maxc = dataSheet.max_column

# Copy Headers
cleanDataSheet.cell(row=1,column=1).value = "Date"
for i in range(1, dataSheet.max_column+1):
    cleanDataSheet.cell(row=1,column=i+1).value = dataSheet.cell(row=1,column=i).value

currDate = date(2022, 3, 1) # Current Date
days = [] # Holds recorded Days
day1 = DayData(currDate) # Create first day

# Check all hourly recordings and store 4 recordings for each day into DayData data structure
for r in range(2, maxr):

    # Extract Hour from HR:MIN:SEC 
    currValue = int(str(dataSheet.cell(row=r,column=1).value)[:2])

    # Check if end of day
    if currValue > int(str(dataSheet.cell(row=r+1,column=1).value)[:2]): 
        currDate+=timedelta(days=1)
    
    # If day is recorded, create new day
    if day1.day != currDate:
        # If last recording of the day is 4th data point, store it
        if day1.fourth == -1:
            day1.fourth = r
            days.append(day1)
        # Create new day
        day1 = DayData(currDate)
        continue

    # Store 1 am, 7 am, 1 pm, and 7 pm hourly recordings from each day
    # If any are unavailable, take the next available hourly recording EX: if 1 am doesnt exist, check for 2 am, 3am, etc..
    if ((currValue == 1 or currValue > 1) and day1.first == -1):
        day1.first = r
    elif ((currValue == 7 or currValue > 7) and day1.second == -1):
        day1.second = r
    elif ((currValue == 13 or currValue > 13) and day1.third == -1):
        day1.third = r
    elif ((currValue == 19 or currValue > 19) and day1.fourth == -1):
        day1.fourth = r
        days.append(day1)

rowInd = 2 # Counter for index of 4 daily hourly recordings
currDate = date(2022, 3, 1) # Reset Current Date variable to first day

# Transfer DayData information into new sheet
for i in range(len(days)):
    for y in [days[i].first, days[i].second, days[i].third,days[i].fourth]:
        cleanDataSheet.cell(row=rowInd, column=1).value = currDate
        for c in range(2, maxc+2):
            cleanDataSheet.cell(row=rowInd, column=c).value = dataSheet.cell(row=y, column=c-1).value
        rowInd+=1   
    currDate+=timedelta(days=1)     

# Save excel document
wb2.save('WeatherNotebook.xlsx')
